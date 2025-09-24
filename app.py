import gradio as gr
import pandas as pd
import json, os, re
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
import openpyxl

replacement_nama_fasilitas = {
    "AEON Credit Services Indonesia": "AEON Credit",
    "Adira Dinamika Multi Finance": "Adira",
    "Akulaku Finance Indonesia": "Akulaku",
    "Atome Finance Indonesia": "Atome Finance",
    "Astra Multi Finance": "Astra MF",
    "BFI Finance Indonesia": "BFI",
    "BIMA Multi Finance": "Bima MF",
    "BPD Jawa Barat dan Banten": "BJB",
    "BPD Jawa Barat dan Banten Syariah": "BJB Syariah",
    "BPD Jawa Timur": "Bank Jatim",
    "BPD Sumatera Utara": "Bank Sumut",
    "Bank BCA Syariah": "BCA Syariah",
    "Bank CIMB Niaga": "CIMB Niaga",
    "Bank Central Asia": "BCA",
    "Bank DBS Indonesia": "Bank DBS",
    "Bank Danamon Indonesia": "Danamon",
    "Bank Danamon Indonesia Syariah": "Danamon Syariah",
    "Bank Hibank Indonesia": "Hibank",
    "Bank HSBC Indonesia": "HSBC",
    "Bank KEB Hana Indonesia": "Bank KEB Hana",
    "Bank Mandiri": "Bank Mandiri",
    "Bank Mandiri Taspen": "Bank Mantap",
    "Bank Mayapada Internasional": "Bank Mayapada",
    "Bank Maybank Indonesia": "Maybank",
    "Bank Mega Syariah": "Bank Mega Syariah",
    "Bank Muamalat Indonesia": "Bank Muamalat",
    "Bank Negara Indonesia": "BNI",
    "Bank Neo Commerce": "Akulaku",
    "Bank OCBC NISP": "OCBC NISP",
    "Bank Panin Indonesia": "Panin Bank",
    "Bank Permata": "Bank Permata",
    "Bank QNB Indonesia": "Bank QNB",
    "Bank Rakyat Indonesia": "BRI",
    "Bank Sahabat Sampoerna": "Bank Sampoerna",
    "Bank Saqu Indonesia (": "Bank Saqu",
    "Bank Seabank Indonesia": "Seabank",
    "Bank SMBC Indonesia": "Bank SMBC",
    "Bank Syariah Indonesia": "BSI",
    "Bank Tabungan Negara": "BTN",
    "Bank UOB Indonesia": "Bank UOB",
    "Bank Woori Saudara": "BWS",
    "Bank Woori Saudara Indonesia 1906":"BWS",
    "Bussan Auto Finance": "BAF",
    "Cakrawala Citra Mega Multifinance":"CCM Finance",
    "Commerce Finance": "Seabank",
    "Dana Mandiri Sejahtera": "Dana Mandiri",
    "Esta Dana Ventura": "Esta Dana",
    "Federal International Finance": "FIF",
    "Globalindo Multi Finance": "Globalindo MF",
    "Home Credit Indonesia": "Home Credit",
    "Indodana Multi Finance": "Indodana MF",
    "Indomobil Finance Indonesia": "IMFI",
    "Indonesia Airawata Finance (": "Indonesia Airawata Finance",
    "JACCS Mitra Pinasthika Mustika Finance Indonesia": "JACCS",
    "KB Finansia Multi Finance": "Kreditplus",
    "Kredivo Finance Indonesia": "Kredivo",
    "Krom Bank Indonesia": "Krom Bank",
    "LOLC Ventura Indonesia": "LOVI",
    "Mandala Multifinance": "Mandala MF",
    "Mandiri Utama Finance": "MUF",
    "Maybank Syariah": "Maybank Syariah",
    "Mega Auto Finance": "MAF",
    "Mega Central Finance": "MCF",
    "Mitra Bisnis Keluarga Ventura": "MBK",
    "Multifinance Anak Bangsa": "MF Anak Bangsa",
    "Panin Bank": "Panin Bank",
    "Permodalan Nasional Madani": "PNM",
    "Pratama Interdana Finance": "Pratama Finance",
    "Standard Chartered Bank": "Standard Chartered",
    "Summit Oto Finance": "Summit Oto",
    "Super Bank Indonesia": "Superbank",
    "Wahana Ottomitra Multiartha": "WOM",
    "Bank Jago": "Bank Jago",
    "Bank BTPN Syariah,": "BTPNS",
    "Bina Artha Ventura": "BAV"
}


def bersihkan_nama_fasilitas(nama_fasilitas: str) -> str:
    if not nama_fasilitas:
        return ""
    lower_fasilitas = nama_fasilitas.lower()
    if "d/h" in lower_fasilitas:
        nama_bersih = nama_fasilitas[:lower_fasilitas.find("d/h")].strip()
    elif "d.h" in lower_fasilitas:
        nama_bersih = nama_fasilitas[:lower_fasilitas.find("d.h")].strip()
    else:
        nama_bersih = nama_fasilitas.strip()
    for pattern in ["PT ", "PT.", "PD.", "(Persero)", "(Perseroda)", "Perseroda", "(UUS)", " Tbk"]:
        nama_bersih = nama_bersih.replace(pattern, "")
    nama_bersih = nama_bersih.replace("Bank Perekonomian Rakyat Syariah", "BPRS")
    nama_bersih = nama_bersih.replace("Bank Perekonomian Rakyat", "BPR")
    nama_bersih = nama_bersih.replace("Koperasi Simpan Pinjam", "KSP")
    nama_bersih = nama_bersih.strip()
    for nama_asli, alias in replacement_nama_fasilitas.items():
        if nama_asli.lower() == nama_bersih.lower():
            return alias
    return nama_bersih


def gabungkan_fasilitas_dengan_jumlah(fasilitas_list):
    counter = Counter(fasilitas_list)
    return '; '.join([f"{nama} ({jumlah})" if jumlah > 1 else nama for nama, jumlah in counter.items()])


def proses_files_debitur(files):
    print("Starting proses_files_debitur")
    if not files:
        print("No files uploaded")
        return pd.DataFrame(), None

    hasil_semua = []
    excluded_fasilitas = {"BTPNS", "Bank Jago", "BAV"}

    for f in files:
        try:
            # Identifikasi nama dan path file
            original_name = (
                getattr(f, "orig_name", None)
                or getattr(f, "name", None)
                or os.path.basename(f.name if hasattr(f, "name") else f)
            )
            path = getattr(f, "name", None) or getattr(f, "path", None) or f
            print(f"Processing file: {original_name}")

            # Hanya proses file .txt
            if not str(original_name).lower().endswith(".txt"):
                print(f"Skipping non-txt file: {original_name}")
                continue

            # Baca file JSON
            try:
                with open(path, "r", encoding="latin-1") as file:
                    data = json.load(file)
                print(f"Successfully read JSON from {original_name}")
            except Exception as e:
                print(f"Gagal membaca file: {original_name} -> {e}")
                continue

            # Struktur TXT
            """
            individual
              dataPokokDebitur
                namaDebitur
              fasilitas
                kreditPembiayan
                  ljkKet
                  bakiDebet
                  tanggalAkadAkhir
                  kualitas
                  kualitasKet
                  jumlahHariTunggakan
                  plafonAwal
                  tunggakanPokok
                  tunggakanBunga
                  denda
                  kondisiKet
                  tahunBulan01Ht
                  tahunBulan01Kol
            """

            # Ambil data pokok debitur
            fasilitas = data.get("individual", {}).get("fasilitas", {}).get("kreditPembiayan", [])
            data_pokok = data.get("individual", {}).get("dataPokokDebitur", [])
            nama_debitur = ", ".join(
                set(
                    debitur.get("namaDebitur", "")
                    for debitur in data_pokok
                    if debitur.get("namaDebitur")
                )
            )
            print(f"Nama Debitur: {nama_debitur}")

            # Definisi awal variabel
            total_plafon = 0
            total_baki_debet = 0
            jumlah_fasilitas_aktif = 0
            kol_1_list, kol_25_list, wo_list, lovi_list, semua_fasilitas_list = [], [], [], [], []
            baki_debet_kol25wo = 0
            baki_debet_wo_all = 0

            # Proses setiap fasilitas
            for item in fasilitas:
                tahun_wo = ""
                kondisi_ket = (item.get("kondisiKet") or "").lower()
                nama_fasilitas = item.get("ljkKet") or ""
                nama_fasilitas_lower = nama_fasilitas.lower()

                baki_debet_val = int(item.get("bakiDebet", 0))
                tunggakan_pokok = int(item.get("tunggakanPokok", 0))
                tunggakan_bunga = int(item.get("tunggakanBunga", 0))
                denda_val = int(item.get("denda", 0))

                nama_fasilitas_bersih = bersihkan_nama_fasilitas(nama_fasilitas)
                semua_fasilitas_list.append(nama_fasilitas_bersih)

                # Tentukan status aktif
                is_aktif = kondisi_ket in ["fasilitas aktif", "diblokir sementara"]
                if not is_aktif and kondisi_ket not in ["lunas", "dihapusbukukan", "hapus tagih"]:
                    if any([baki_debet_val > 0, tunggakan_pokok > 0, tunggakan_bunga > 0, denda_val > 0]):
                        is_aktif = True

                # Skip lunas kecuali LOVI
                if kondisi_ket == "lunas" and "pt lolc ventura indonesia" not in nama_fasilitas_lower:
                    continue

                jumlah_hari_tunggakan = int(item.get("jumlahHariTunggakan", 0))
                kualitas = item.get("kualitas", "")
                kol_value = f"{kualitas}/{jumlah_hari_tunggakan}" if jumlah_hari_tunggakan != 0 else kualitas
                tanggal_kondisi = item.get("tanggalKondisi", "")
                baki_debet = baki_debet_val

                # Pada WO, jika baki debet jika nol cek tunggakan pokok dsb
                if kondisi_ket in ["dihapusbukukan", "hapus tagih"] or is_aktif:
                    if baki_debet == 0:
                        baki_debet = tunggakan_pokok + tunggakan_bunga + denda_val
                        if baki_debet == 0:
                            kondisi_ket = "lunas"
                            is_aktif = False

                plafon_awal = int(item.get("plafonAwal", 0))
                nama_fasilitas_bersih = bersihkan_nama_fasilitas(nama_fasilitas)
                baki_debet_format = "{:,.0f}".format(baki_debet).replace(",", ".")

                # Penulisan fasilitas dengan kolektibilitas
                if is_aktif and kualitas == "1" and jumlah_hari_tunggakan <= 30:
                    fasilitas_teks = nama_fasilitas_bersih
                elif is_aktif:
                    fasilitas_teks = f"{nama_fasilitas_bersih} Kol {kol_value} {baki_debet_format}"
                elif kondisi_ket in ["dihapusbukukan", "hapus tagih"]:
                    try:
                        tahun_wo = int(str(tanggal_kondisi)[:4])
                    except:
                        tahun_wo = ""
                    fasilitas_teks = f"{nama_fasilitas_bersih} WO {tahun_wo} {baki_debet_format}"
                else:
                    fasilitas_teks = nama_fasilitas_bersih

                # Penulisan LOVI dengan kolektibilitas
                if kondisi_ket == "lunas":
                    fasilitas_lovi = "Lunas"
                elif is_aktif:
                    fasilitas_lovi = f"Kol {kol_value}"
                elif kondisi_ket in ["dihapusbukukan", "hapus tagih"]:
                    fasilitas_lovi = f"WO {tahun_wo} {baki_debet_format}"
                else:
                    fasilitas_lovi = nama_fasilitas_bersih

                # Penulisan fasilitas dan dipisahkan kolom
                if "pt lolc ventura indonesia" not in nama_fasilitas_lower:
                    if is_aktif:
                        total_plafon += plafon_awal
                        total_baki_debet += baki_debet
                        jumlah_fasilitas_aktif += 1
                        if kualitas == "1" and jumlah_hari_tunggakan <= 30:
                            if jumlah_hari_tunggakan == 0:
                                kol_1_list.append(nama_fasilitas_bersih)
                            else:
                                kol_1_list.append(f"{nama_fasilitas_bersih} Kol {kualitas}/{jumlah_hari_tunggakan}")
                        else:
                            kol_25_list.append(fasilitas_teks)
                            if nama_fasilitas_bersih not in excluded_fasilitas:
                                baki_debet_kol25wo += baki_debet
                    elif kondisi_ket in ["dihapusbukukan", "hapus tagih"]:
                        wo_list.append(fasilitas_teks)
                        if nama_fasilitas_bersih not in excluded_fasilitas:
                            baki_debet_kol25wo += baki_debet
                        baki_debet_wo_all += baki_debet
                else:
                    if is_aktif or kondisi_ket in ["lunas", "dihapusbukukan"]:
                        tanggal_akad_akhir = item.get("tanggalAkadAkhir", "")
                        if tanggal_akad_akhir:
                            if not lovi_list:
                                lovi_list.append({"keterangan": fasilitas_lovi, "tanggal": tanggal_akad_akhir})
                            elif tanggal_akad_akhir > lovi_list[0]["tanggal"]:
                                lovi_list[0] = {"keterangan": fasilitas_lovi, "tanggal": tanggal_akad_akhir}



            # Penulisan satu data terburuk
            kol_terburuk = None

            for f in fasilitas:
                kondisi_ket = (f.get("kondisiKet") or "").lower()

                if kondisi_ket in ["dihapusbukukan", "hapus tagih"]:
                    continue

                for i in range(1, 25):
                    kol_key = f"tahunBulan{i:02d}Kol"
                    if kol_key in f:
                        try:
                            kol_val = int(f.get(kol_key) or 0)
                        except:
                            continue

                        if kol_val > 0:
                            if kol_terburuk is None or kol_val > kol_terburuk:
                                kol_terburuk = kol_val

            if kol_terburuk is None:
                for f in fasilitas:
                    for i in range(1, 25):
                        kol_key = f"tahunBulan{i:02d}Kol"
                        if kol_key in f:
                            try:
                                kol_val = int(f.get(kol_key) or 0)
                            except:
                                continue

                            if kol_val > 0:
                                if kol_terburuk is None or kol_val > kol_terburuk:
                                    kol_terburuk = kol_val

            kol_terburuk_str = f"Kol {kol_terburuk}" if kol_terburuk else ""

            # Logika Rekomendasi
            if jumlah_fasilitas_aktif >= 0 and not kol_25_list and not wo_list and not lovi_list:
                rekomendasi = "OK"
            elif any("lunas" in lovi.get("keterangan", "").lower() or "kol 1" in lovi.get("keterangan", "").lower() for lovi in lovi_list):
                rekomendasi = "OK"
            elif jumlah_fasilitas_aktif >= 0 and baki_debet_kol25wo <= 250_000 and not lovi_list:
                rekomendasi = "OK"
            else:
                rekomendasi = "NOT OK"

            # Simpan hasil per file
            filename = os.path.basename(original_name or path)
            nik = os.path.splitext(filename)[0]

            # Update baca nama file NIK
            for prefix in ["NIK_", "KTP_", "Paspor_"]:
                if nik.upper().startswith(prefix.upper()):
                    nik = nik[len(prefix):]

            hasil_semua.append({
                "NIK": nik,
                "Nama Debitur": nama_debitur,
                "Rekomendasi": rekomendasi,
                "Jumlah Fasilitas": jumlah_fasilitas_aktif,
                "Total Plafon Awal": total_plafon if jumlah_fasilitas_aktif > 0 else "",
                "Total Baki Debet": total_baki_debet if jumlah_fasilitas_aktif > 0 else "",
                "Kol 1": gabungkan_fasilitas_dengan_jumlah(kol_1_list),
                "Kol 2-5": "; ".join(kol_25_list),
                "WO/dihapusbukukan": "; ".join(wo_list),
                "LOVI": "; ".join([l.get("keterangan", "") for l in lovi_list]),
                "Baki Debet Kol25WO": baki_debet_kol25wo,
                "Semua Fasilitas": gabungkan_fasilitas_dengan_jumlah(semua_fasilitas_list),
                "Kol Terburuk": kol_terburuk_str,
                "Baki Debet Kol 1-5": total_baki_debet if jumlah_fasilitas_aktif > 0 else 0,
                "Baki Debet WO All": baki_debet_wo_all,

            })
            print(f"Finished processing file: {original_name}")
        except Exception as e:
            print(f"Error processing file {original_name}: {e}")


    if not hasil_semua:
        print("No valid files processed")
        return pd.DataFrame(), None

    # Penyatuan Data dengan NIK >= 2 data
    grouped = defaultdict(list)
    for row in hasil_semua:
        nik_raw = str(row["NIK"]).strip()

        # Update baca nama file NIK
        for prefix in ["NIK_", "KTP_", "Paspor_"]:
            if nik_raw.upper().startswith(prefix.upper()):
                nik_raw = nik_raw[len(prefix):]

        nik_key = nik_raw.split("-")[0]

        grouped[nik_key].append(row)

    def gabung_kolom(key, is_numerik=False):
        if is_numerik:
            return sum(row[key] for row in rows if isinstance(row[key], (int, float)))
        gabungan = "; ".join(str(row[key]) for row in rows if row[key])
        return "; ".join(sorted(set(gabungan.split("; "))))

    hasil_digabung = []
    for nik_key, rows in grouped.items():
        if len(rows) == 1:
            hasil_digabung.append(rows[0])
        else:
            hasil_digabung.append({
                "NIK": str(nik_key),
                "Nama Debitur": gabung_kolom("Nama Debitur"),
                "Rekomendasi": gabung_kolom("Rekomendasi"),
                "Jumlah Fasilitas": gabung_kolom("Jumlah Fasilitas", is_numerik=True),
                "Total Plafon Awal": gabung_kolom("Total Plafon Awal", is_numerik=True),
                "Total Baki Debet": gabung_kolom("Total Baki Debet", is_numerik=True),
                "Kol 1": gabung_kolom("Kol 1"),
                "Kol 2-5": gabung_kolom("Kol 2-5"),
                "WO/dihapusbukukan": gabung_kolom("WO/dihapusbukukan"),
                "LOVI": gabung_kolom("LOVI"),
                "Baki Debet Kol25WO": gabung_kolom("Baki Debet Kol25WO", is_numerik=True),
                "Semua Fasilitas": gabung_kolom("Semua Fasilitas"),
                "Kol Terburuk": gabung_kolom("Kol Terburuk"),
                "Baki Debet Kol 1-5": gabung_kolom("Baki Debet Kol 1-5", is_numerik=True),
                "Baki Debet WO All": gabung_kolom("Baki Debet WO All", is_numerik=True)

            })

    # Logika Rekomendasi setelah Penggabungan
    hasil_final = []
    for row in hasil_digabung:
        jumlah_fasilitas_aktif = row["Jumlah Fasilitas"]
        baki_debet_kol25wo = row.get("Baki Debet Kol25WO", 0)

        kol_25_list = row["Kol 2-5"].split("; ") if row["Kol 2-5"] else []
        wo_list = row["WO/dihapusbukukan"].split("; ") if row["WO/dihapusbukukan"] else []
        lovi_list = row["LOVI"].split("; ") if row["LOVI"] else []

        if jumlah_fasilitas_aktif >= 0 and not kol_25_list and not wo_list and not lovi_list:
            rekomendasi = "OK"
        elif any("lunas" in lovi.lower() or "kol 1" in lovi.lower() for lovi in lovi_list):
            rekomendasi = "OK"
        elif jumlah_fasilitas_aktif >= 0 and baki_debet_kol25wo <= 250_000 and not lovi_list:
            rekomendasi = "OK"
        else:
            rekomendasi = "NOT OK"

        row["Rekomendasi"] = rekomendasi
        hasil_final.append(row)

    # Logika Score
    for row in hasil_final:
        jumlah_fasilitas_aktif = row.get("Jumlah Fasilitas", 0) or 0
        semua_fasilitas_val = str(row.get("Semua Fasilitas", "")).strip().lower()
        if re.match(r"^lovi(\s*\(\d+\))?$", semua_fasilitas_val):
            semua_fasilitas_val = "lovi"
        kol_terburuk_val = str(row.get("Kol Terburuk", "")).strip().lower()
        lovi_raw = str(row.get("LOVI", "")).strip().lower()
        match = re.search(r"kol\s*\d+", lovi_raw)
        lovi_val = match.group(0) if match else lovi_raw
        baki_debet_kol15 = row.get("Baki Debet Kol 1-5", 0) or 0
        baki_debet_wo_all = row.get("Baki Debet WO All", 0) or 0

        score = None
        if (
            semua_fasilitas_val != ""
            and (kol_terburuk_val == "" or kol_terburuk_val == "kol 1")
            and jumlah_fasilitas_aktif == 0
            and baki_debet_wo_all == 0
            and (lovi_val == "" or lovi_val == "lunas")
        ):
            score = 0
        elif (
            (
                jumlah_fasilitas_aktif > 0
                or (
                    jumlah_fasilitas_aktif == 0
                    and semua_fasilitas_val != "lovi"
                    and lovi_val == "kol 1"
                    and baki_debet_kol15 == 0
                )
            )
            and kol_terburuk_val == "kol 1"
            and baki_debet_kol15 <= 10_000_000
            and baki_debet_wo_all == 0
        ):
            score = 1
        elif (
            semua_fasilitas_val == ""
            or (semua_fasilitas_val == "lovi" and lovi_val == "kol 1")
            or (
                (
                    jumlah_fasilitas_aktif == 0
                    or
                    jumlah_fasilitas_aktif > 0
                    or (
                        jumlah_fasilitas_aktif == 0
                        and semua_fasilitas_val != "lovi"
                        and lovi_val == "kol 1"
                        and baki_debet_kol15 == 0
                    )
                )
                and (
                    kol_terburuk_val == "kol 1"
                    or (
                        kol_terburuk_val == "kol 1"
                        and baki_debet_wo_all >= 0
                    )
                    )
            )
        ):
            score = 2
        elif (
              semua_fasilitas_val != ""
              and kol_terburuk_val != "kol 1"
              and jumlah_fasilitas_aktif == 0
              and baki_debet_wo_all >= 0
              and (lovi_val == "" or lovi_val == "lunas")
        ):
            score = 3
        elif (
            (
                jumlah_fasilitas_aktif > 0
                or (
                    jumlah_fasilitas_aktif == 0
                    and lovi_val != ""
                    and lovi_val != "lunas"
                    and baki_debet_kol15 == 0
                )
            )
            and baki_debet_kol15 <= 10_000_000
            and kol_terburuk_val != "kol 1"
        ):
            score = 4
        elif (
            (
                jumlah_fasilitas_aktif > 0
                or (
                    jumlah_fasilitas_aktif == 0
                    and lovi_val != ""
                    and lovi_val != "lunas"
                    and baki_debet_kol15 == 0
                )
            )
            and baki_debet_kol15 > 10_000_000
            and kol_terburuk_val != "kol 1"
        ):
            score = 5

        row["Score"] = score

    # Output Excel
    df = pd.DataFrame(hasil_final)

    # Sembunyikan kolom di excels
    hapus_kolom = ["Baki Debet Kol25WO", "Semua Fasilitas", "Kol Terburuk", "Baki Debet Kol 1-5", "Baki Debet WO All"]
    df.drop(columns=hapus_kolom, inplace=True, errors="ignore")
    df.sort_values(by="NIK", inplace=True)
    tanggal_hari_ini = datetime.today().strftime("%d-%m-%Y_%H%M%S")
    output_file = f"Hasil SLIK Debitur {tanggal_hari_ini}.xlsx"
    df.to_excel(output_file, index=False)
    print(f"Excel file created: {output_file}")

    # Format Excel
    custom_widths = {
        "NIK": 17,
        "Nama Debitur": 22,
        "Rekomendasi": 12,
        "Jumlah Fasilitas": 8,
        "Total Plafon Awal": 12,
        "Total Baki Debet": 12,
        "Kol 1": 30,
        "Kol 2-5": 30,
        "WO/dihapusbukukan": 30,
        "LOVI": 9,
        "Semua Fasilitas": 25,
        "Kol Terburuk": 9,
        "Baki Debet Kol 1-5": 11,
        "Baki Debet WO All": 11,
        "Score": 5
    }
    wrap_columns = set(custom_widths.keys())
    center_columns = set(custom_widths.keys())
    number_format_columns = {"Total Plafon Awal", "Total Baki Debet", "Baki Debet Kol 1-5", "Baki Debet WO All"}

    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        header_row = ws[1]
        header = [cell.value for cell in header_row]

        for idx, col_cells in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(idx)
            col_name = header[idx - 1] if idx - 1 < len(header) else ""

            wrap = col_name in wrap_columns
            center = col_name in center_columns

            if center and wrap:
                alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif center:
                alignment = Alignment(horizontal="center", vertical="center")
            elif wrap:
                alignment = Alignment(wrap_text=True)
            else:
                alignment = Alignment()

            for i, cell in enumerate(col_cells):
                cell.alignment = alignment
                cell.font = Font(size=8)
                cell.border = thin_border

                if col_name == "NIK":
                    cell.number_format = "@"
                elif i != 0 and col_name in number_format_columns:
                    cell.number_format = "#,##0"

            if col_name in custom_widths:
                ws.column_dimensions[col_letter].width = custom_widths[col_name]
            else:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
                ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(output_file)
        print("Excel formatting applied")
    except Exception as e:
        print(f"Error formatting Excel file: {e}")
        # Return the dataframe and file path even if formatting fails
        return df, output_file


    print("Finished proses_files_debitur successfully")
    return df, output_file

def proses_files_karyawan(files):
    print("Starting proses_files_karyawan")
    if not files:
        print("No files uploaded")
        return pd.DataFrame(), None

    hasil_semua = []

    for f in files:
        try:
            original_name = getattr(f, "orig_name", None) or getattr(f, "name", None) or os.path.basename(f.name if hasattr(f, "name") else f)
            path = getattr(f, "name", None) or getattr(f, "path", None) or f
            print(f"Processing file: {original_name}")

            if not str(original_name).lower().endswith(".txt"):
                print(f"Skipping non-txt file: {original_name}")
                continue

            try:
                with open(path, "r", encoding="latin-1") as file:
                    data = json.load(file)
                print(f"Successfully read JSON from {original_name}")
            except Exception as e:
                print(f"Gagal membaca file: {original_name} -> {e}")
                continue

            fasilitas = data.get('individual', {}).get('fasilitas', {}).get('kreditPembiayan', [])
            data_pokok = data.get('individual', {}).get('dataPokokDebitur', [])
            nama_debitur = ', '.join(set(debitur.get('namaDebitur', '') for debitur in data_pokok if debitur.get('namaDebitur')))
            print(f"Nama Karyawan: {nama_debitur}")

            total_plafon = 0
            total_baki_debet = 0
            jumlah_fasilitas_aktif = 0
            kol_1_list, kol_2_list, kol_3_list, kol_4_list, kol_5_list, wo_list = [], [], [], [], [], []

            for item in fasilitas:
                kondisi_ket = (item.get('kondisiKet') or '').lower()
                nama_fasilitas = item.get('ljkKet') or ''
                nama_fasilitas_lower = nama_fasilitas.lower()

                baki_debet_val = int(item.get('bakiDebet', 0))
                tunggakan_pokok = int(item.get('tunggakanPokok', 0))
                tunggakan_bunga = int(item.get('tunggakanBunga', 0))
                denda_val = int(item.get('denda', 0))

                is_aktif = kondisi_ket in ['fasilitas aktif', 'diblokir sementara']

                if not is_aktif and kondisi_ket not in ['lunas', 'dihapusbukukan', 'hapus tagih']:
                    if any([baki_debet_val > 0, tunggakan_pokok > 0, tunggakan_bunga > 0, denda_val > 0]):
                        is_aktif = True

                jumlah_hari_tunggakan = int(item.get('jumlahHariTunggakan', 0))
                kualitas = item.get('kualitas', '')
                kol_value = f"{kualitas}/{jumlah_hari_tunggakan}"
                tanggal_kondisi = item.get('tanggalKondisi', '')
                baki_debet = baki_debet_val

                if kondisi_ket in ["dihapusbukukan", "hapus tagih"] or is_aktif:
                    if baki_debet == 0:
                        baki_debet = tunggakan_pokok + tunggakan_bunga + denda_val
                        if baki_debet == 0:
                            kondisi_ket = "lunas"
                            is_aktif = False

                plafon_awal = int(item.get('plafonAwal', 0))
                nama_fasilitas_bersih = bersihkan_nama_fasilitas(nama_fasilitas)
                baki_debet_format = "{:,.0f}".format(baki_debet).replace(",", ".")
                fasilitas_teks = f"{nama_fasilitas_bersih} Kol {kol_value} {baki_debet_format}"

                if is_aktif:
                    if kualitas == '1':
                        if jumlah_hari_tunggakan <= 30:
                            kol_1_list.append(fasilitas_teks)
                        else:
                            kol_2_list.append(fasilitas_teks)
                    elif kualitas == '2':
                        kol_2_list.append(fasilitas_teks)
                    elif kualitas == '3':
                        kol_3_list.append(fasilitas_teks)
                    elif kualitas == '4':
                        kol_4_list.append(fasilitas_teks)
                    elif kualitas == '5':
                        kol_5_list.append(fasilitas_teks)

                    total_plafon += plafon_awal
                    total_baki_debet += baki_debet
                    jumlah_fasilitas_aktif += 1

                elif kondisi_ket in ['dihapusbukukan', 'hapus tagih']:
                    wo_list.append(f"{nama_fasilitas_bersih} WO {tanggal_kondisi[:4]} {baki_debet_format}")

            print(f"Finished processing file: {original_name}")
            filename = os.path.basename(original_name or path)
            nik = os.path.splitext(filename)[0]

            # Update baca nama file NIK
            for prefix in ["NIK_", "KTP_", "Paspor_"]:
                if nik.upper().startswith(prefix):
                    nik = nik[len(prefix):]

            hasil_semua.append({
                'NIK': nik,
                'Nama Karyawan': nama_debitur,
                'Jumlah Fasilitas': jumlah_fasilitas_aktif,
                'Total Plafon Awal': total_plafon if jumlah_fasilitas_aktif > 0 else "",
                'Total Baki Debet': total_baki_debet if jumlah_fasilitas_aktif > 0 else "",
                'Kol 1': gabungkan_fasilitas_dengan_jumlah(kol_1_list),
                'Kol 2': '; '.join(kol_2_list),
                'Kol 3': '; '.join(kol_3_list),
                'Kol 4': '; '.join(kol_4_list),
                'Kol 5': '; '.join(kol_5_list),
                'WO/dihapusbukukan': '; '.join(wo_list)
            })
        except Exception as e:
            print(f"Error processing file {original_name}: {e}")

        grouped = defaultdict(list)
        for row in hasil_semua:
            nik_raw = str(row["NIK"]).strip()
    
            # Update baca nama file NIK
            for prefix in ["NIK_", "KTP_", "Paspor_"]:
                if nik_raw.upper().startswith(prefix.upper()):
                    nik_raw = nik_raw[len(prefix):]
    
            nik_key = nik_raw.split("-")[0]

            grouped[nik_key].append(row)

    hasil_digabung = []
    for nik_key, rows in grouped.items():
        if len(rows) == 1:
            hasil_digabung.append(rows[0])
            continue

        def gabung_kolom(key, is_numerik=False):
            if is_numerik:
                return sum(row[key] for row in rows if isinstance(row[key], (int, float)))
            gabungan = '; '.join(str(row[key]) for row in rows if row[key])
            return '; '.join(sorted(set(gabungan.split('; '))))

        hasil_digabung.append({
            'NIK': nik_key,
            'Nama Karyawan': gabung_kolom('Nama Karyawan'),
            'Jumlah Fasilitas': gabung_kolom('Jumlah Fasilitas', is_numerik=True),
            'Total Plafon Awal': gabung_kolom('Total Plafon Awal', is_numerik=True),
            'Total Baki Debet': gabung_kolom('Total Baki Debet', is_numerik=True),
            'Kol 1': gabung_kolom('Kol 1'),
            'Kol 2': gabung_kolom('Kol 2'),
            'Kol 3': gabung_kolom('Kol 3'),
            'Kol 4': gabung_kolom('Kol 4'),
            'Kol 5': gabung_kolom('Kol 5'),
            'WO/dihapusbukukan': gabung_kolom('WO/dihapusbukukan')
        })

    df = pd.DataFrame(hasil_digabung)
    df.sort_values(by='NIK', inplace=True)
    output_file = f'Hasil SLIK Karyawan {datetime.today().strftime("%d-%m-%Y_%H%M%S")}.xlsx'
    df.to_excel(output_file, index=False)
    print(f"Excel file created: {output_file}")


    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        custom_widths = {
            'NIK': 16,
            'Nama Karyawan': 20,
            'Jumlah Fasilitas': 8,
            'Total Plafon Awal': 11,
            'Total Baki Debet': 11,
            'Kol 1': 20,
            'Kol 2': 20,
            'Kol 3': 20,
            'Kol 4': 20,
            'Kol 5': 20,
            'WO/dihapusbukukan': 20
        }
        wrap_columns = set(custom_widths.keys())
        center_columns = set(custom_widths.keys())
        number_format_columns = {'Total Plafon Awal', 'Total Baki Debet'}
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        header_row = ws[1]
        header = [cell.value for cell in header_row]
        for idx, col_cells in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(idx)
            col_name = header[idx - 1] if idx - 1 < len(header) else ''
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=col_name in wrap_columns)
            for i, cell in enumerate(col_cells):
                cell.alignment = alignment
                cell.font = Font(size=8)
                cell.border = thin_border

                if col_name == "NIK":
                    cell.number_format = "@"
                elif i != 0 and col_name in number_format_columns:
                    cell.number_format = '#,##0'

            if col_name in custom_widths:
                ws.column_dimensions[col_letter].width = custom_widths[col_name]

        wb.save(output_file)
        print("Excel formatting applied")
    except Exception as e:
        print(f"Error formatting Excel file: {e}")
        # Return the dataframe and file path even if formatting fails
        return df, output_file

    print("Finished proses_files_karyawan successfully")
    return df, output_file

def clear_data():
    return None, None, pd.DataFrame()


# UI Graduio
with gr.Blocks(
    theme=gr.themes.Soft(),
    css="""
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;500;600&display=swap');

        * {
            font-family: 'Poppins', sans-serif !important;
        }

        h3 {
            text-align: center !important;
            margin-top: 60px !important;
            font-weight: normal !important;
            font-size: 12pt !important;
        }

        /* Navbar */
        .navbar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 10px 20px;
            border-bottom: 1px solid #ddd;
            position: sticky;
            top: 0;
            z-index: 100;
        }

        .navbar-title {
            font-weight: 600;
            font-size: 20px;
        }
        /* Button Row */
        .center-row {
            display: flex !important;
            justify-content: center !important;
            gap: 20px !important;
        }

        .image-button {
            width: 150px !important;
            height: 150px !important;
            min-width: 150px !important;
            max-width: 150px !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            padding: 0 !important;
            border-radius: 10px !important;
            font-size: 14px !important;
            text-align: center !important;
            white-space: normal !important;
            word-break: break-word !important;
        }

        .image-button:hover {
            transform: scale(1.05);
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }

        .small-text,
        .small-text * {
            font-size: 10px !important;
            color: #555 !important;
        }

        #preview-table,
        #preview-table * {
            font-size: 10px !important;
        }

        .tight-row {
            gap: 20px !important;
            justify-content: center;
            flex-wrap: wrap;
        }

        footer {
            text-align: center;
            margin-top: 10px;
            color: gray;
            font-size: 13px;
        }
    """
) as demo:


    # Halaman Pilihan
    with gr.Column(visible=True) as halaman_pilihan:
        with gr.Row(elem_classes="navbar"):
            navbar_title = gr.Markdown("**SLIK Data Processing System**", elem_classes="navbar-title")

        gr.Markdown("<h3 style='margin-bottom: 30px;'>Select data type to continue</h3>")
        with gr.Row(elem_classes="center-row"):
            tombol_debitur = gr.Button("Debtor Data", elem_classes="image-button")
            tombol_karyawan = gr.Button("Employee Data", elem_classes="image-button")

        gr.HTML("<footer style='margin-top: 60px !important; '>© 2025 | Created by Ayu Nurhasanah</footer>")

    # Halaman Debitur
    with gr.Column(visible=False) as halaman_debitur:
        with gr.Row(elem_classes="navbar"):
            navbar_title_debitur = gr.Markdown("**Debtor Data Processing**", elem_classes="navbar-title")

        tombol_kembali_debitur = gr.Button("Back", size="sm", scale=0, min_width=100)
        gr.Markdown("Upload several .txt files, then click **Process**, then download the processed Excel file.")

        with gr.Row():
            with gr.Column(scale=1):
                inp_files_debitur = gr.File(label="", file_count="multiple", file_types=[".txt"])
                tombol_proses_debitur = gr.Button("Process", variant="primary")
            with gr.Column(scale=1):
                output_file_debitur = gr.File(label="Download", file_types=[".xlsx"])
                clear_btn_debitur = gr.Button("Clear Data", variant="secondary")

        output_df_debitur = gr.Dataframe(label="Preview", elem_id="preview-table", wrap=False)
        gr.HTML("<footer>© 2025 | Created by Ayu Nurhasanah</footer>")

    # Halaman Karyawan
    with gr.Column(visible=False) as halaman_karyawan:
        with gr.Row(elem_classes="navbar"):
            navbar_title_karyawan = gr.Markdown("**Employee Data Processing**", elem_classes="navbar-title")

        tombol_kembali_karyawan = gr.Button("Back", size="sm", scale=0, min_width=100)
        gr.Markdown("Upload several .txt files, then click **Process**, then download the processed Excel file.")

        with gr.Row():
            with gr.Column(scale=1):
                inp_files_karyawan = gr.File(label="", file_count="multiple", file_types=[".txt"])
                tombol_proses_karyawan = gr.Button("Process", variant="primary")
            with gr.Column(scale=1):
                output_file_karyawan = gr.File(label="Download", file_types=[".xlsx"])
                clear_btn_karyawan = gr.Button("Clear Data", variant="secondary")

        output_df_karyawan = gr.Dataframe(label="Preview", elem_id="preview-table", wrap=False)
        gr.HTML("<footer>© 2025 | Created by Ayu Nurhasanah</footer>")


    # Navigasi Tombol
    tombol_debitur.click(lambda: (gr.update(visible=False), gr.update(visible=True), gr.update(visible=False)),
                         outputs=[halaman_pilihan, halaman_debitur, halaman_karyawan])
    tombol_karyawan.click(lambda: (gr.update(visible=False), gr.update(visible=False), gr.update(visible=True)),
                          outputs=[halaman_pilihan, halaman_debitur, halaman_karyawan])
    tombol_kembali_debitur.click(lambda: (gr.update(visible=True), gr.update(visible=False), gr.update(visible=False)),
                                 outputs=[halaman_pilihan, halaman_debitur, halaman_karyawan])
    tombol_kembali_karyawan.click(lambda: (gr.update(visible=True), gr.update(visible=False), gr.update(visible=False)),
                                  outputs=[halaman_pilihan, halaman_debitur, halaman_karyawan])


    # Proses Data
    tombol_proses_debitur.click(fn=proses_files_debitur, inputs=[inp_files_debitur],
                                outputs=[output_df_debitur, output_file_debitur])
    clear_btn_debitur.click(fn=clear_data, outputs=[inp_files_debitur, output_file_debitur, output_df_debitur])

    tombol_proses_karyawan.click(fn=proses_files_karyawan, inputs=[inp_files_karyawan],
                                 outputs=[output_df_karyawan, output_file_karyawan])
    clear_btn_karyawan.click(fn=clear_data, outputs=[inp_files_karyawan, output_file_karyawan, output_df_karyawan])


demo.launch(
    server_name="0.0.0.0",
    server_port=int(os.environ.get("PORT", 7860))
)
